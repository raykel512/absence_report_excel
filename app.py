import streamlit as st
import pandas as pd
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

# openpyxl에서 인쇄 방향 상수를 직접 import
# from openpyxl.worksheet.page import PageSetup # 주석 처리하고 아래 문자열 사용

st.set_page_config(page_title="자동 결석 신고서 생성기 (Excel)", layout="centered")
st.title("📝 자동 결석 신고서 생성 (Excel 형식)")
st.caption("A4 용지 한 페이지에 인쇄되도록 최적화된 Excel 파일을 생성합니다.")

# ----------------------------------------------------
# A. 데이터 입력값 설정
# ----------------------------------------------------

# 예시 학생 명단 (선택 박스용)
STUDENTS = {
    "10101": {"학년": 1, "반": 1, "번호": 1, "이름": "김철수"},
    "10102": {"학년": 1, "반": 1, "번호": 2, "이름": "이영희"},
    "20315": {"학년": 2, "반": 3, "번호": 15, "이름": "박민재"},
}

st.subheader("1. 결석 학생 정보 입력")
student_options = {f"{s['학년']}-{s['반']}-{s['번호']} {s['이름']}": k for k, s in STUDENTS.items()}
selected_key = st.selectbox(
    "학생 선택",
    options=list(student_options.keys()),
    index=None
)

if selected_key:
    student_data = STUDENTS[student_options[selected_key]]
    
    # 총 일수 계산
    def calculate_days(start, end):
        if start > end: return 0
        return (end - start).days + 1
        
    st.subheader("2. 결석 기간 및 사유")
    
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("시작일", date.today())
    with col2:
        end_date = st.date_input("종료일", date.today())
    
    total_days = calculate_days(start_date, end_date)
    st.markdown(f"**👉 총 결석 예상 일수 (단순 계산): {total_days}일**")
        
    reason = st.text_area("결석 사유", "독감으로 인한 자가 격리")
    
    st.subheader("3. 결석 종류 및 첨부 서류 정보")
    absence_type = st.radio(
        "결석 종류 선택",
        options=['질병', '인정', '기타'],
        index=0
    )
    
    col_chk1, col_chk2 = st.columns(2)
    with col_chk1:
        has_diagnosis = st.checkbox("진단서/진료확인서 첨부 (3일 이상인 경우)", value=(total_days >= 3 and absence_type == '질병'))
    with col_chk2:
        has_opinion = st.checkbox("보건결석 학부모 의견서 첨부 (보건 결석인 경우)", value=(absence_type == '인정'))
        
    etc_doc_val = st.text_input("기타 첨부 서류 명칭", "")
    
    # ----------------------------------------------------
    # B. Excel 생성 및 PDF 양식 서식 적용 함수
    # ----------------------------------------------------
    
    def create_excel_report(data, has_diagnosis, has_opinion, etc_doc_val):
        wb = Workbook()
        ws = wb.active
        ws.title = "결석신고서"
        
        # --- 서식 정의 ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        title_font = Font(size=14, bold=True)
        header_font = Font(bold=True)
        
        # A4 너비에 맞게 E열까지만 사용하도록 열 너비 조정
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # --- 1. 문서 제목 및 안내 ---
        current_row = 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "학업성적관리규정 [결석계 서식]"
        ws[f'A{current_row}'].font = Font(size=10)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "결 석 신 고 서"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_align
        ws.row_dimensions[current_row].height = 25
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "※ 결석신고서는 결석한 날로부터 3일 이내에 제출하여 학교의 승인을 받아야 합니다."
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 15

        # --- 2. 학생 정보 ---
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "학생"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'] = f"{data['학년']}학년 {data['반']}반 {data['번호']}번"
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border
        
        # --- 3. 기간 ---
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "기간"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border
        
        period_str = f"{data['시작일'].strftime('2025년 %m월 %d일')}부터 ~ {data['종료일'].strftime('2025년 %m월 %d일')}까지 ({data['총_일수']}일간)"
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'] = period_str
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 20
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "※ 결석 기간 중 공휴일 또는 학교 휴무일은 결석일 수에 포함하지 않습니다."
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 15

        # --- 4. 성명 ---
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "성명"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'] = data['이름']
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border
        
        # --- 5. 사유 ---
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "사유"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'] = data['사유']
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 60 
        
        # --- 6. 붙임 서류 ---
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "붙임 서류"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border
        
        doc_list = []
        doc_list.append(f"[{'X' if has_diagnosis else ' '}] 진단서 또는 진료 확인서 (3일 이상인 경우 꼭 첨부)")
        doc_list.append(f"[] 병원처방전 또는 약봉투") 
        doc_list.append(f"[{'X' if has_opinion else ' '}] 보건결석 학부모 의견서")
        
        is_none = not (has_diagnosis or has_opinion or etc_doc_val.strip())
        doc_list.append(f"[{'X' if is_none else ' '}] 없음")
        
        if etc_doc_val.strip():
            doc_list.append(f"[{'X'}] 기타 ({etc_doc_val})")
        else:
             doc_list.append(f"[] 기타 ()")

        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'] = '\n'.join(doc_list)
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 70
        
        # --- 7. 유의사항 및 보호자 연서 ---
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "※ 규정된 증빙서류를 첨부하지 않으면 '미인정(무단)' 결석 처리됩니다."
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws.row_dimensions[current_row].height = 15

        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = f"위와 같이 결석하고자 하였기에 보호자 연서로 신고합니다. \n\n {date.today().strftime('2025년 %m월 %d일')}"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='bottom', wrap_text=True)
        ws.row_dimensions[current_row].height = 40
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = f"학생 성명: {data['이름']} (서명 또는 인)"
        ws.merge_cells(f'D{current_row}:E{current_row}')
        ws[f'D{current_row}'] = "보호자 성명: (서명 또는 인)"
        ws[f'A{current_row}'].alignment = left_align
        ws[f'D{current_row}'].alignment = left_align
        ws.row_dimensions[current_row].height = 30
        
        # --- 8. 담임교사 확인서 (새로운 섹션) ---
        current_row += 2
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "담임교사 확인서"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_align
        ws.row_dimensions[current_row].height = 25
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "결석 종류"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border
        
        chk_질병 = 'X' if data['결석_종류'] == '질병' else ' '
        chk_인정 = 'X' if data['결석_종류'] == '인정' else ' '
        chk_기타 = 'X' if data['결석_종류'] == '기타' else ' '
        
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'] = f"[{chk_질병}] 질병  [{chk_인정}] 인정  [{chk_기타}] 기타"
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border
        
        # 확인 방법 
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "확인 방법"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'C{current_row}:E{current_row}')
        ws[f'C{current_row}'] = "[X] 제출된 증빙서류로 확인"
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border
        
        # --- 9. 교사 확인 텍스트 및 날짜 ---
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "위의 신고 내용이 사실과 같음을 확인합니다." # 누락된 텍스트 추가
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 20

        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = f"{date.today().strftime('2025년 %m월 %d일')}"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='bottom')
        ws.row_dimensions[current_row].height = 25


        # --- 10. 결재 라인 ---
        current_row += 1
        
        # 결재 라인 헤더
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = "학급 담임"
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].alignment = center_align
        
        ws[f'C{current_row}'] = "출결 담당"
        ws[f'C{current_row}'].border = thin_border
        ws[f'C{current_row}'].alignment = center_align
        
        ws[f'D{current_row}'] = "교무 부장"
        ws[f'D{current_row}'].border = thin_border
        ws[f'D{current_row}'].alignment = center_align
        
        ws[f'E{current_row}'] = "교감"
        ws[f'E{current_row}'].border = thin_border
        ws[f'E{current_row}'].alignment = center_align
        
        # 최종 서명/결재 빈칸 (공간 확보)
        current_row += 1
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thin_border
            ws.row_dimensions[current_row].height = 30
        
        # 학교장 귀하
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "대동세무고등학교장 귀하"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[current_row].height = 20
        
        # --- 11. 2페이지 내용 (규정 상세) 추가 ---
        current_row += 2
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "※ 결석 종류별 증빙자료 관련 규정 안내 (PDF 2페이지 내용)"
        ws[f'A{current_row}'].font = Font(size=10, bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        ws[f'A{current_row}'].alignment = left_align
        ws.row_dimensions[current_row].height = 20
        
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        rule_text = (
            "1. 질병결석 2일 이내: 결석신고서와 담임교사 확인서\n"
            "2. 질병결석 3일 이상: 결석신고서, 담임교사 확인서 및 ① 의사의 진단서, ② 의견서(진료확인서 등) 중 택1\n"
            "3. 보건결석: 의사소견서 또는 학부모 의견서 첨부 (월 1일만 인정)\n"
            "4. 그 외 인정 및 기타결석: 사유를 인정할 수 있는 증빙서류 첨부\n"
            "5. 고사기간 중의 질병결석: 의사의 진단서 반드시 첨부"
        )
        ws[f'A{current_row}'] = rule_text
        ws[f'A{current_row}'].font = Font(size=9)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 80
        
        # 인쇄 영역 설정 (A4 1페이지에 맞춤)
        ws.page_setup.fitToPages = True
        ws.page_setup.fitToWidth = 1 # 너비를 1페이지에 맞춤
        ws.page_setup.fitToHeight = 0 # 높이는 맞추지 않음 (1페이지를 넘을 경우 다음 페이지로 넘김)
        
        # 🌟 오류 수정: 상수 대신 문자열 'portrait' 사용
        ws.page_setup.orientation = 'portrait' 
        
        ws.print_area = f'A1:E{current_row}'

        return wb

    # ----------------------------------------------------
    # C. 파일 생성 및 다운로드 (이전과 동일)
    # ----------------------------------------------------
    
    # 최종 대체 데이터 조합
    final_data = {
        "학년": student_data["학년"], "반": student_data["반"], "번호": student_data["번호"],
        "이름": student_data["이름"], "총_일수": total_days,
        "시작일": start_date, "종료일": end_date,
        "사유": reason, "결석_종류": absence_type
    }

    st.markdown("---")
    if st.button("결석 신고서 생성 및 다운로드 (Excel)", use_container_width=True):
        st.subheader("4. 결과 확인")
        
        # Excel 문서 생성
        workbook = create_excel_report(final_data, has_diagnosis, has_opinion, etc_doc_val)
        
        # BytesIO를 사용하여 메모리에 문서를 저장하고 Streamlit 다운로드에 사용
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        file_name = f"결석신고서_Excel_{final_data['이름']}_{final_data['시작일'].strftime('%Y%m%d')}.xlsx"
        
        st.download_button(
            label=f"📥 {file_name} 다운로드",
            data=excel_buffer,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.success("Excel 신고서 생성이 완료되었습니다! 다운로드 후 인쇄하여 사용하세요.")
        st.balloons()

else:
    st.info("먼저 결석한 학생을 선택해주세요.")
